"""
JobTracker SaaS - Routes d'export
"""

from fastapi import APIRouter, Depends, Response
from fastapi.responses import StreamingResponse
from typing import List
from datetime import datetime, timezone
import json
import csv
import io

from utils.auth import get_current_user

router = APIRouter(prefix="/export", tags=["Export"])


def get_db():
    """Dependency injection pour la DB"""
    pass


@router.get("/json")
async def export_json(
    current_user: dict = Depends(get_current_user),
    db = Depends(get_db)
):
    """Export toutes les candidatures en JSON"""
    user_id = current_user["user_id"]
    
    # Récupérer les candidatures
    applications = await db.applications.find(
        {"user_id": user_id},
        {"_id": 0, "user_id": 0}
    ).sort("date_candidature", -1).to_list(10000)
    
    # Récupérer les entretiens et les associer
    for app in applications:
        interviews = await db.interviews.find(
            {"candidature_id": app["id"]},
            {"_id": 0, "user_id": 0, "candidature_id": 0}
        ).to_list(100)
        app["entretiens"] = interviews
    
    # Créer le JSON
    export_data = {
        "export_date": datetime.now(timezone.utc).isoformat(),
        "total_applications": len(applications),
        "applications": applications
    }
    
    json_str = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
    
    return Response(
        content=json_str,
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=candidatures_{datetime.now().strftime('%Y%m%d')}.json"
        }
    )


@router.get("/csv")
async def export_csv(
    current_user: dict = Depends(get_current_user),
    db = Depends(get_db)
):
    """Export toutes les candidatures en CSV"""
    user_id = current_user["user_id"]
    
    # Récupérer les candidatures
    applications = await db.applications.find(
        {"user_id": user_id},
        {"_id": 0, "user_id": 0}
    ).sort("date_candidature", -1).to_list(10000)
    
    # Créer le CSV
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_ALL)
    
    # En-têtes
    headers = [
        "Entreprise", "Poste", "Type", "Lieu", "Moyen",
        "Date Candidature", "Lien", "Statut", "Date Réponse",
        "Commentaire", "Favori"
    ]
    writer.writerow(headers)
    
    # Données
    for app in applications:
        row = [
            app.get("entreprise", ""),
            app.get("poste", ""),
            app.get("type_poste", ""),
            app.get("lieu", ""),
            app.get("moyen", ""),
            app.get("date_candidature", "")[:10] if app.get("date_candidature") else "",
            app.get("lien", ""),
            app.get("reponse", ""),
            app.get("date_reponse", "")[:10] if app.get("date_reponse") else "",
            app.get("commentaire", ""),
            "Oui" if app.get("is_favorite") else "Non"
        ]
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=candidatures_{datetime.now().strftime('%Y%m%d')}.csv"
        }
    )


@router.get("/excel")
async def export_excel(
    current_user: dict = Depends(get_current_user),
    db = Depends(get_db)
):
    """Export toutes les candidatures en Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(
            content=json.dumps({"error": "openpyxl non installé"}),
            media_type="application/json",
            status_code=500
        )
    
    user_id = current_user["user_id"]
    
    # Récupérer les candidatures
    applications = await db.applications.find(
        {"user_id": user_id},
        {"_id": 0}
    ).sort("date_candidature", -1).to_list(10000)
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidatures"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    gold_fill = PatternFill(start_color="c4a052", end_color="c4a052", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = [
        "Entreprise", "Poste", "Type", "Lieu", "Moyen",
        "Date Candidature", "Lien", "Statut", "Date Réponse",
        "Commentaire", "Favori", "Nb Entretiens"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Données
    for row_idx, app in enumerate(applications, 2):
        # Compter les entretiens
        interviews_count = await db.interviews.count_documents({"candidature_id": app["id"]})
        
        data = [
            app.get("entreprise", ""),
            app.get("poste", ""),
            app.get("type_poste", ""),
            app.get("lieu", ""),
            app.get("moyen", ""),
            app.get("date_candidature", "")[:10] if app.get("date_candidature") else "",
            app.get("lien", ""),
            app.get("reponse", ""),
            app.get("date_reponse", "")[:10] if app.get("date_reponse") else "",
            app.get("commentaire", ""),
            "⭐" if app.get("is_favorite") else "",
            interviews_count
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            
            # Colorer les favoris
            if app.get("is_favorite"):
                ws.cell(row=row_idx, column=11).fill = gold_fill
    
    # Ajuster largeurs
    column_widths = [25, 30, 12, 15, 20, 15, 50, 15, 15, 40, 10, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Figer la première ligne
    ws.freeze_panes = "A2"
    
    # Sauvegarder en mémoire
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=candidatures_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )


@router.get("/statistics/excel")
async def export_statistics_excel(
    current_user: dict = Depends(get_current_user),
    db = Depends(get_db)
):
    """Export des statistiques en Excel multi-sheets"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response(
            content=json.dumps({"error": "openpyxl non installé"}),
            media_type="application/json",
            status_code=500
        )
    
    user_id = current_user["user_id"]
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Stats générales
    ws1 = wb.active
    ws1.title = "Stats Générales"
    
    total = await db.applications.count_documents({"user_id": user_id})
    pending = await db.applications.count_documents({"user_id": user_id, "reponse": "pending"})
    positive = await db.applications.count_documents({"user_id": user_id, "reponse": "positive"})
    negative = await db.applications.count_documents({"user_id": user_id, "reponse": "negative"})
    
    stats = [
        ("Total Candidatures", total),
        ("En attente", pending),
        ("Réponses positives", positive),
        ("Réponses négatives", negative),
        ("Taux de réponse", f"{((positive + negative) / total * 100):.1f}%" if total > 0 else "0%")
    ]
    
    for row, (label, value) in enumerate(stats, 1):
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=2, value=value)
    
    # Sheet 2: Par statut
    ws2 = wb.create_sheet("Par Statut")
    pipeline = [
        {"$match": {"user_id": user_id}},
        {"$group": {"_id": "$reponse", "count": {"$sum": 1}}}
    ]
    status_data = await db.applications.aggregate(pipeline).to_list(10)
    
    ws2.cell(row=1, column=1, value="Statut")
    ws2.cell(row=1, column=2, value="Nombre")
    for row, item in enumerate(status_data, 2):
        ws2.cell(row=row, column=1, value=item["_id"] or "pending")
        ws2.cell(row=row, column=2, value=item["count"])
    
    # Sheet 3: Par type
    ws3 = wb.create_sheet("Par Type")
    pipeline = [
        {"$match": {"user_id": user_id}},
        {"$group": {"_id": "$type_poste", "count": {"$sum": 1}}}
    ]
    type_data = await db.applications.aggregate(pipeline).to_list(10)
    
    ws3.cell(row=1, column=1, value="Type")
    ws3.cell(row=1, column=2, value="Nombre")
    for row, item in enumerate(type_data, 2):
        ws3.cell(row=row, column=1, value=item["_id"] or "cdi")
        ws3.cell(row=row, column=2, value=item["count"])
    
    # Sauvegarder
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=statistiques_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )


# ============== EXPORT ENTRETIENS ==============

@router.get("/interviews/json")
async def export_interviews_json(
    current_user: dict = Depends(get_current_user),
    db = Depends(get_db)
):
    """Export tous les entretiens en JSON"""
    user_id = current_user["user_id"]
    
    # Récupérer les entretiens avec les infos de candidature
    interviews = await db.interviews.find(
        {"user_id": user_id},
        {"_id": 0, "user_id": 0}
    ).sort("date_entretien", -1).to_list(10000)
    
    # Enrichir avec les infos de candidature
    for interview in interviews:
        if interview.get("candidature_id"):
            app = await db.applications.find_one(
                {"id": interview["candidature_id"]},
                {"_id": 0, "entreprise": 1, "poste": 1}
            )
            if app:
                interview["entreprise"] = app.get("entreprise", "")
                interview["poste"] = app.get("poste", "")
    
    export_data = {
        "export_date": datetime.now(timezone.utc).isoformat(),
        "total_interviews": len(interviews),
        "interviews": interviews
    }
    
    json_str = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
    
    return Response(
        content=json_str,
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=entretiens_{datetime.now().strftime('%Y%m%d')}.json"
        }
    )


@router.get("/interviews/csv")
async def export_interviews_csv(
    current_user: dict = Depends(get_current_user),
    db = Depends(get_db)
):
    """Export tous les entretiens en CSV"""
    user_id = current_user["user_id"]
    
    interviews = await db.interviews.find(
        {"user_id": user_id},
        {"_id": 0}
    ).sort("date_entretien", -1).to_list(10000)
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_ALL)
    
    headers = [
        "Entreprise", "Poste", "Date Entretien", "Type", "Format",
        "Lieu/Lien", "Recruteur", "Statut", "Commentaire"
    ]
    writer.writerow(headers)
    
    for interview in interviews:
        # Get application info
        entreprise = ""
        poste = ""
        if interview.get("candidature_id"):
            app = await db.applications.find_one(
                {"id": interview["candidature_id"]},
                {"_id": 0, "entreprise": 1, "poste": 1}
            )
            if app:
                entreprise = app.get("entreprise", "")
                poste = app.get("poste", "")
        
        row = [
            entreprise,
            poste,
            interview.get("date_entretien", "")[:16] if interview.get("date_entretien") else "",
            interview.get("type_entretien", ""),
            interview.get("format_entretien", ""),
            interview.get("lieu_lien", ""),
            interview.get("interviewer", ""),
            interview.get("statut", ""),
            interview.get("commentaire", "")
        ]
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=entretiens_{datetime.now().strftime('%Y%m%d')}.csv"
        }
    )


@router.get("/interviews/excel")
async def export_interviews_excel(
    current_user: dict = Depends(get_current_user),
    db = Depends(get_db)
):
    """Export tous les entretiens en Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(
            content=json.dumps({"error": "openpyxl non installé"}),
            media_type="application/json",
            status_code=500
        )
    
    user_id = current_user["user_id"]
    
    interviews = await db.interviews.find(
        {"user_id": user_id},
        {"_id": 0}
    ).sort("date_entretien", -1).to_list(10000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entretiens"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = [
        "Entreprise", "Poste", "Date Entretien", "Type", "Format",
        "Lieu/Lien", "Recruteur", "Statut", "Commentaire"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_idx, interview in enumerate(interviews, 2):
        # Get application info
        entreprise = ""
        poste = ""
        if interview.get("candidature_id"):
            app = await db.applications.find_one(
                {"id": interview["candidature_id"]},
                {"_id": 0, "entreprise": 1, "poste": 1}
            )
            if app:
                entreprise = app.get("entreprise", "")
                poste = app.get("poste", "")
        
        data = [
            entreprise,
            poste,
            interview.get("date_entretien", "")[:16] if interview.get("date_entretien") else "",
            interview.get("type_entretien", ""),
            interview.get("format_entretien", ""),
            interview.get("lieu_lien", ""),
            interview.get("interviewer", ""),
            interview.get("statut", ""),
            interview.get("commentaire", "")
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Adjust column widths
    column_widths = [25, 30, 18, 15, 12, 40, 20, 15, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=entretiens_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )
